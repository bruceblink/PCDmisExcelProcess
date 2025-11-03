import sys
import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

def log(msg):
    """输出日志并写入文件"""
    print(msg)
    with open("process.log", "a", encoding="utf-8") as f:
        f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}\n")

def get_values(ws, col, max_rows=20):
    """读取一列数据"""
    vals = []
    for i in range(1, max_rows + 1):
        vals.append(ws[f"{col}{i}"].value)
    return vals

def safe_float(s, default=None):
    try:
        return float(s)
    except (ValueError, TypeError):
        return default

def start(target_path: str, origin_path: str):
    log(f"=== 开始执行 ===")
    log(f"源文件: {origin_path}")
    log(f"目标文件: {target_path}")

    if not os.path.exists(origin_path):
        log(f"❌ 找不到源文件: {origin_path}")
        return

    if not os.path.exists(target_path):
        log(f"❌ 找不到目标文件: {target_path}")
        return

    wb_target = load_workbook(target_path)
    wb_origin = load_workbook(origin_path, data_only=True)

    # === 1. 启动条件检查 ===
    if "Sheet1" not in wb_target.sheetnames:
        log("❌ 目标文件中没有 Sheet1")
        return

    ws_sheet1 = wb_target["Sheet1"]
    # check_value = ws_sheet1["F29"].value

    # if check_value not in ["李春宁", "刘文"]:
    #     log("⚠️ 启动条件不满足：Sheet1!F29 不是 '李春宁' 或 '刘文'")
    #     return
    # else:
    #     log(f"✅ 启动条件通过: {check_value}")

    # === 2. 获取源文件的报告表数据 ===
    if "PCDmisExcel1" not in wb_origin.sheetnames:
        log("❌ 报告文件中找不到 PCDmisExcel1 工作表")
        return

    ws_report = wb_origin["PCDmisExcel1"]
    log("✅ 找到 PCDmisExcel1")

    # === 3. 读取基础数据 C/F/G/D/A ===
    dataC = get_values(ws_report, "C")
    dataF = get_values(ws_report, "F")
    dataG = get_values(ws_report, "G")
    dataD = get_values(ws_report, "D")
    dataA = get_values(ws_report, "A")
    # 定义写入目标文件的数据
    arr_target = [["" for _ in range(5)] for _ in range(20)]
    for i in range(20):
        c, f, g, d, a = dataC[i], dataF[i], dataG[i], dataD[i], dataA[i]
        if c is None and f is None and g is None and d is None and a is None:
            continue
        arr_target[i][0] = c
        arr_target[i][2] = f
        arr_target[i][3] = g
        # 将D列的字符串安全抓换成数字
        d_num = safe_float(d, default=0.0)
        arr_target[i][1] = a if (d_num == 0) else d  # 默认情况下取d列的值，如果D列为0则取a列
        if g not in (None, ""):
            arr_target[i][4] = "CMM"

    # === 4. 写入 A8:E27 ===
    for ws in wb_target.worksheets:
        for r in range(8, 28):
            for c in range(1, 6):
                ws.cell(r, c, None)
        for r in range(20):
            for c in range(5):
                ws.cell(r + 8, c + 1, arr_target[r][c])

    log("✅ 写入 A8:E27 完成")

    # === 5. 收集源文件中的 PCDmisExcel 工作表 ===
    pcd_sheets = [s for s in wb_origin.sheetnames if s.startswith("PCDmisExcel")]
    pcd_sheets = pcd_sheets[:200]
    log(f"源文件中共找到 {len(pcd_sheets)} 个 PCDmisExcel 工作表")

    pcd_data = {}

    for idx, sheet_name in enumerate(pcd_sheets):
        ws = wb_origin[sheet_name]

        def get_last_row(col):
            for row in range(20, 0, -1):
                if ws[f"{col}{row}"].value not in (None, ""):
                    return row
            return 0

        row_h = get_last_row("H")
        row_i = get_last_row("I")
        row_count = max(row_h, row_i, 0)
        if row_count == 0:
            continue

        # 判断使用 H 还是 I 列
        sumH = 0.0
        for r in range(1, 21):
            val = ws[f"H{r}"].value
            try:
                if val not in (None, ""):
                    sumH += float(val)
            except Exception as e:
                log(e)
                pass

        data_col = "I" if sumH == 0 else "H"
        data_vals = [ws[f"{data_col}{r}"].value for r in range(1, row_count + 1)]

        pcd_data[sheet_name] = data_vals
        log(f"读取 {sheet_name}: {len(data_vals)} 行, 使用列 {data_col}")

    # === 6. 写入目标文件 F8:Y27 并填充红色 ===
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    for ws in wb_target.worksheets:
        for r in range(8, 28):
            for c in range(6, 26):
                ws.cell(r, c, None)
                ws.cell(r, c).fill = PatternFill()  # 清空原有填充

    for i, (sheet_name, values) in enumerate(pcd_data.items()):
        backup_index = i // 20
        backup_col_offset = (i % 20) + 6
        if backup_index < len(wb_target.worksheets):
            ws_target = wb_target.worksheets[backup_index]
            ws_source = wb_origin[sheet_name]
            for r, val in enumerate(values[:20]):
                target_cell = ws_target.cell(r + 8, backup_col_offset, val)

                # 获取源文件对应的 F/G/I 列值， 转换成float
                f_val = safe_float(ws_source[f"F{r+1}"].value, default=0.0) # 正公差
                g_val = safe_float(ws_source[f"G{r+1}"].value, default=0.0) # 负公差
                i_val = safe_float(ws_source[f"I{r+1}"].value, default=0.0) # 检查值
                try:
                    if i_val is not None:
                        # 大于正公差 或者小于负公差，说明超出公差范围，填充标记为红色
                        if (g_val is not None and i_val > f_val) or (f_val is not None and i_val < g_val):
                            target_cell.fill = red_fill
                except Exception as e:
                    log(e)
                    pass

    log("✅ 写入 F8:Y27 完成并应用红色填充")

    # === 7. 删除空白工作表 ===
    sheets_to_delete = []
    for ws in wb_target.worksheets:
        if ws["F8"].value in (None, "", "###EMPTY###"):
            sheets_to_delete.append(ws.title)

    if len(sheets_to_delete) < len(wb_target.worksheets):
        for name in sheets_to_delete:
            del wb_target[name]
        log(f"🗑️ 删除空白工作表: {sheets_to_delete}")
    else:
        log("⚠️ 所有工作表的F8都为空，至少保留一个工作表！")

    # === 8. 更新 检验日期 ===
    ws_sheet1['C4'] = datetime.now().strftime("%Y.%m.%d")

    # === 9. 保存结果 ===
    wb_target.save(target_path)
    log(f"✅ 数据处理完成！结果保存在：{target_path}")
    log("=== 执行结束 ===\n")


if __name__ == "__main__":
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()

    origin_file = filedialog.askopenfilename(title="请选择源文件", filetypes=[("Excel 文件", "*.xlsx")])
    if not origin_file:
        messagebox.showwarning("提示", "未选择源文件，已取消。")
        sys.exit()

    template_file = filedialog.askopenfilename(title="请选择模板文件 template.xlsx", filetypes=[("Excel 文件", "*.xlsx")])
    if not template_file:
        messagebox.showwarning("提示", "未选择模板文件，已取消。")
        sys.exit()

    # === ✅ 在当前运行目录生成目标文件 ===
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target_file = os.path.join(
        os.getcwd(),  # 当前程序运行目录
        f"output_{timestamp}.xlsx"
    )
    shutil.copyfile(template_file, target_file)
    log(f"📂 已复制模板文件为新目标文件: {target_file}")

    # === 在新文件上执行处理 ===
    start(target_file, origin_file)

    messagebox.showinfo("完成", f"Excel 数据处理完成！\n结果文件：\n{target_file}\n程序运行的详细信息见 process.log。")
